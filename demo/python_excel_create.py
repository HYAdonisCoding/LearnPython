#/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
#
#实例工作簿
wb = Workbook()

#通过工作簿 激活 默认工作表
apple_ws = wb.active

#为工作表 修改 名字为：苹果营业额
apple_ws.title = "苹果营业额"

#通过工作簿 创建 工作簿 指定其名字为：樱桃营业额
cherry_ws = wb.create_sheet("樱桃营业额")

#通过工作簿 创建 工作簿 指定其名字为：数据核算表
check_ws = wb.create_sheet("数据核算表")

#构建 苹果营业额 樱桃营业额 表头
xl_head = ["季度", "成本", "销售额", "单价"]

#构建 苹果营业额表格数据 注意：是二维的数据，因为表格有行有列
apple_data = [xl_head, ["Q1", 2000, 5000, 5], ["Q2", 3000, 6000, 6], ["Q3", 4000, 9000, 9],["Q4", 4000, 3000, 6],
              ["Q1", 2000, 5000, 5], ["Q2", 3000, 6000, 6], ["Q3", 4000, 9000, 9],["Q4", 4000, 3000, 6],
              ["Q1", 2000, 5000, 5], ["Q2", 3000, 6000, 6], ["Q3", 4000, 9000, 9],["Q4", 4000, 3000, 6]]
#构建 樱桃营业额表格数据
cherry_data = [xl_head, ["Q1", 5000, 10000, 25], ["Q2", 6000, 8000, 40], ["Q3", 4000, 9000, 30],["Q4", 6000, 12000, 20]]

#构建 数据核算表表格数据
check_data = [["品种", "季度"], ["苹果", "Q1"], ["苹果", "Q2"], ["苹果", "Q3"], ["苹果","Q4"], ["车厘子", "Q1"], ["车厘子", "Q2"], ["车厘子", "Q3"], ["车厘子","Q4"]]

#将数据写入到工作表中
for row_data in apple_data:
    apple_ws.append(row_data)


for row_data in cherry_data:
    cherry_ws.append(row_data)

for row_data in check_data:
    check_ws.append(row_data)

#保存工作簿  注意：openpyxl 不支持.xls
wb.save("老Amy水果店营业额.xlsx")