#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
#
# #实例工作簿
# wb = Workbook()
#
# #通过工作簿 激活 默认工作表
# apple_ws = wb.active
#
# #为工作表 修改 名字为：苹果营业额
# apple_ws.title = "苹果营业额"
#
# #通过工作簿 创建 工作簿 指定其名字为：樱桃营业额
# cherry_ws = wb.create_sheet("樱桃营业额")
#
# #通过工作簿 创建 工作簿 指定其名字为：数据核算表
# check_ws = wb.create_sheet("数据核算表")
#
# #构建 苹果营业额 樱桃营业额 表头
# xl_head = ["季度", "成本", "销售额", "单价"]
#
# #构建 苹果营业额表格数据 注意：是二维的数据，因为表格有行有列
# apple_data = [xl_head, ["Q1", 2000, 5000, 5], ["Q2", 3000, 6000, 6], ["Q3", 4000, 9000, 9],["Q4", 4000, 3000, 6]]
#
# #构建 樱桃营业额表格数据
# cherry_data = [xl_head, ["Q1", 5000, 10000, 25], ["Q2", 6000, 8000, 40], ["Q3", 4000, 9000, 30],["Q4", 6000, 12000, 20]]
#
# #构建 数据核算表表格数据
# check_data = [["品种", "季度"], ["苹果", "Q1"], ["苹果", "Q2"], ["苹果", "Q3"], ["苹果","Q4"], ["车厘子", "Q1"], ["车厘子", "Q2"], ["车厘子", "Q3"], ["车厘子","Q4"]]
#
# #将数据写入到工作表中
# for row_data in apple_data:
#     apple_ws.append(row_data)
#
# for row_data in cherry_data:
#     cherry_ws.append(row_data)
#
# for row_data in check_data:
#     check_ws.append(row_data)
#
# #保存工作簿  注意：openpyxl 不支持.xls
# wb.save("老Amy水果店营业额.xlsx")
#
# #2根据苹果营业额与车厘子营业额两张表，为数据核算表计算添加 销售数量、利润两列
import pandas as pd
import numpy as np

excel = "老Amy水果店营业额.xlsx"
sheet_name = "苹果营业额"
apple_data = pd.read_excel(excel, sheet_name=sheet_name)
# apple_data1 = pd.read_excel("老Amy水果店营业额.xlsx", sheet_name="苹果营业额", index_col=3)
print('apple_data1\n', apple_data)
# for apple in apple_data:
#     print('apple', apple)
apple_data.drop_duplicates()
print('apple_data\n', apple_data.values.size, apple_data.values)


def checkSame(target, a):
    b = []
    for index, nums in enumerate(a):
        print('index, nums, target', index, nums, target)
        if (nums == target).all():
            b.append(index)
    print('不', b)
    return b


# cherry_data = pd.read_excel("老Amy水果店营业额.xlsx",sheet_name="樱桃营业额")
# print(cherry_data)
#
# check_data = pd.read_excel("老Amy水果店营业额.xlsx",sheet_name="数据核算表")
# print(check_data)
#
# check_data["销售数量"] = np.nan
# check_data["利润"] = np.nan
# print(check_data)
# #
# # #计算苹果每季度的销售数量与利润
# apple_data["销售数量"] = apple_data["销售额"]/apple_data["单价"]
# apple_data["利润"] = apple_data["销售额"] - apple_data["成本"]
# print(apple_data)
# #
# cherry_data["销售数量"] = cherry_data["销售额"]/cherry_data["单价"]
# cherry_data["利润"] = cherry_data["销售额"] - cherry_data["成本"]
# print(cherry_data)
#
# #使 苹果营业额表 与 车厘子营业额表 进行堆叠
# df = apple_data.append(cherry_data)
# print(df)
# #
# # # 获取 df 数据中的销售数量与利润两列赋值给 check_data 对应两列即可
# check_data["销售数量"] = df["销售数量"].values
# check_data["利润"] = df["利润"].values
# print(check_data)
#
# # 需求3
#
# # 将苹果营业额与车厘子营业额根据 季度 横向连接
# # 计算每季度水果
# # 总销售额
# # 总成本
# # 总利润
#
# # 代码魔法
# # 根据 on 指定公共列 横向连接
# ver_data = cherry_data.merge(apple_data, on="季度",suffixes=("_cherry", "_apple"))
# print(ver_data)
#
# # 计算每季度水果的总指标
# ver_data["季度总成本"] = ver_data["成本_cherry"] + ver_data["成本_apple"]
# ver_data["季度总利润"] = ver_data["利润_cherry"] + ver_data["利润_apple"]
# print('ver_data', ver_data)


# 删除

# 引入openpyxl模块
import openpyxl


def delete_rows(rows, excel, sheet_name):
    # 让wb = 工作薄，已有的工作薄路径
    wb = openpyxl.load_workbook(excel)

    # 让ws等于工作薄中的Sheet1工作表
    ws = wb[sheet_name]
    # 第一个无需删除
    rows.remove(rows[0])
    print('rows', reversed(rows))
    # 删除从第3行开始算的2行内容
    for index in reversed(rows):
        print(index)
        ws.delete_rows(index + 2, 1)
    # 删除从第1列开始算的2列内容
    # ws.delete_cols(1,2)

    # 另存为表格
    wb.save(excel)


# same = checkSame(apple_data.values[1], apple_data.values)
# delete_rows(same, excel, sheet_name)

def checkAll():
    apple_data_o = pd.read_excel(excel, sheet_name=sheet_name)
    for values in apple_data_o.values:
        print('values', values)
        same = checkSame(values, apple_data_o.values)
        print('same', same)
        delete_rows(same, excel, sheet_name)
        apple_data_o = pd.read_excel(excel, sheet_name=sheet_name)
    print('finished')


checkAll()
